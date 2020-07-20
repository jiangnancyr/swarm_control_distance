import xlrd
import xlwt
import openpyxl
fnr = 'initDatas/文件1.xlsx'
fnw = 'initDatas/文件1.xlsx'

'''
规定excel表的第一行不要写数据
读excel几个比较重要的函数：
re = ged.readExcel(fileName=fileName) #读文件
re.nrows #文件有多少行
rowDatas = re.readRows(i)#读取一行数据
we = ged.writeExcel(fileName=saveFile) #写文件句柄
we.writeRows(rowDatas1,k) #写一行数据到指定行
we.saveExcel() #写文件要保存文件

'''

#设置表格样式
def set_style(name,height,bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    return style    

class readExcel:

    def __init__(self, fileName=fnr):

        self.wb = xlrd.open_workbook(filename=fileName)#打开文件
        # print(self.wb.sheet_names())#获取所有表格名字
        # self.nrows = self.sheet1.nrows
        # self.ncols = self.sheet1.ncols

    def readRows(self,
                 n,
                 name=''):
        self.sheet = self.wb.sheet_by_name(name)  # 通过索引获取表格
        rows = self.sheet.row_values(n)#获取行内容
        return rows
    
    def readCols(self,
                 n,
                 name=''):
        self.sheet = self.wb.sheet_by_name(name)  # 通过索引获取表格
        cols = self.sheet.col_values(n)#获取列内容
        return cols

    def readCell (self,
                r,
                c,
                name=''):
        self.sheet = self.wb.sheet_by_name(name)  # 通过索引获取表格
        return self.sheet.cell_value(r - 1, c - 1)
    
class writeExcel:
    def __init__(self,fileName=fnw):
        try:
            re = readExcel(fileName=fileName)
            sheetName = re.wb.sheet_names()
            self.outWb = openpyxl.load_workbook(fileName)  # 打开一个将写的文件
            self.sheet = self.outWb.get_sheet_by_name(sheetName[0])

        except FileNotFoundError: 
            self.outWb = openpyxl.Workbook()  # 打开一个将写的文件
            sheetName = self.outWb.get_sheet_names()
            self.sheet = self.outWb.get_sheet_by_name(sheetName[0])
        self.saveName = fileName

    def getSheetByName(self,
                       name):
        try:
            self.sheet = self.outWb.get_sheet_by_name(name)
        except KeyError:
            self.sheet = self.outWb.create_sheet(name)
        return self.sheet

    def writeRows(self,
                  n,
                  data,
                  name='',):
        self.getSheetByName(name)
        for i in range(len(data)):
            self.sheet.cell(n, i + 1).value = data[i]

    def writeCell(self,
                  n,
                  z,
                  data,
                  title = '',):
        self.getSheetByName(title)
        self.sheet.cell(n, z).value = data

    def saveExcel(self):
        self.outWb.save(self.saveName)
        
def main ():
    we = writeExcel(fileName='../resData/tests.xlsx')
    we.writeCell(1,1,3,'syncRate')
    we.writeRows(1, [1,2,3,4], 'syncRate')
    we.saveExcel()
    we.writeCell(1, 1, 3, 'syncRate')
    we.writeRows(2, [4, 3, 2, 1], 'syncRate')
    we.saveExcel()
    re = readExcel(fileName='../resData/tests.xlsx')
    res = re.readCell(2, 2,'syncRate')

    print(res)
if __name__ ==  '__main__':
    main ()
    
